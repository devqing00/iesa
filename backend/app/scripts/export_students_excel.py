"""
Export Students to Excel

Generates an Excel workbook with two sheets:
  Sheet 1 - All Students: Full details of every registered user
  Sheet 2 - Dues Paid:    Only students who have paid the basic due (₦4,000)
                          in the currently active academic session.

Usage:
    python -m app.scripts.export_students_excel
    python -m app.scripts.export_students_excel --output my_report.xlsx
"""

import asyncio
import argparse
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

# ── DNS Override (must happen BEFORE pymongo/motor is imported) ───────────────
# On networks with restricted internal DNS servers (e.g. school/office/VPN)
# the mongodb+srv:// SRV lookup fails. Patch dnspython to use Google public DNS.
try:
    import dns.resolver
    _public_resolver = dns.resolver.Resolver(configure=False)
    _public_resolver.nameservers = ["8.8.8.8", "8.8.4.4", "1.1.1.1"]
    dns.resolver.default_resolver = _public_resolver
except Exception:
    pass  # If dns isn't available yet, motor will handle the error itself
# ─────────────────────────────────────────────────────────────────────────────

from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient

# openpyxl is used for Excel generation (already likely in the venv; install if missing)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl is not installed. Run:  pip install openpyxl")
    sys.exit(1)

# ── Environment ──────────────────────────────────────────────────────────────
load_dotenv()
MONGODB_URL  = os.getenv("MONGODB_URL",  "mongodb://localhost:27017")
DATABASE_NAME = os.getenv("DATABASE_NAME", "iesa_db")

# ── Styling helpers ──────────────────────────────────────────────────────────

HEADER_FILL_ALL   = PatternFill("solid", fgColor="1F3864")   # dark navy   – all-students sheet
HEADER_FILL_PAID  = PatternFill("solid", fgColor="1A6634")   # dark green  – dues-paid sheet
HEADER_FONT       = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL          = PatternFill("solid", fgColor="EEF2FF")   # soft lavender for alternating rows
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def _header_fill(fill: PatternFill, ws, headers: list[str]):
    """Write styled header row to a worksheet."""
    ws.row_dimensions[1].height = 22
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill   = fill
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _write_row(ws, row_idx: int, values: list, alternate: bool = False):
    """Write a data row with optional alternating row shading."""
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.border    = THIN_BORDER
        if alternate:
            cell.fill = ALT_FILL


def _autofit_columns(ws, headers: list[str], min_width: int = 12, max_width: int = 40):
    """Auto-size column widths based on header length and content."""
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        # Measure header + a few content rows
        max_len = len(str(header))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_width), max_width)


# ── Data helpers ─────────────────────────────────────────────────────────────

def _fmt_date(val) -> str:
    """Safely format a datetime / date / None."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M")
    return str(val)


def _fmt_list(val) -> str:
    if not val:
        return ""
    return ", ".join(str(v) for v in val)


def _flatten_user(user: dict) -> list:
    """Return a flat list of field values for the ALL-STUDENTS sheet."""
    return [
        str(user.get("_id", "")),
        user.get("firstName", ""),
        user.get("lastName", ""),
        user.get("matricNumber") or "",
        user.get("email", ""),
        user.get("secondaryEmail") or "",
        user.get("phone") or "",
        user.get("role", ""),
        user.get("department", ""),
        str(user.get("admissionYear") or ""),
        user.get("currentLevel") or "",
        user.get("gender") or "",
        str(user.get("dateOfBirth") or ""),
        _fmt_list(user.get("skills")),
        user.get("bio") or "",
        "Yes" if user.get("emailVerified") else "No",
        "Yes" if user.get("hasCompletedOnboarding") else "No",
        "Yes" if user.get("isExternalStudent") else "No",
        "Yes" if user.get("isActive", True) else "No",
        _fmt_date(user.get("createdAt")),
        _fmt_date(user.get("lastLogin")),
    ]


ALL_HEADERS = [
    "User ID", "First Name", "Last Name", "Matric Number",
    "Primary Email", "Secondary Email", "Phone",
    "Role", "Department", "Admission Year", "Current Level",
    "Gender", "Date of Birth", "Skills", "Bio",
    "Email Verified", "Onboarding Done", "External Student",
    "Active", "Created At", "Last Login",
]

PAID_HEADERS = [
    "User ID", "First Name", "Last Name", "Matric Number",
    "Primary Email", "Phone",
    "Role", "Department", "Admission Year", "Current Level",
    "Gender", "Active", "Created At",
    "Payment Title", "Amount (₦)", "Payment Date",
]


# ── Core export logic ─────────────────────────────────────────────────────────

async def run_export(output_path: str):
    client = AsyncIOMotorClient(MONGODB_URL)
    db = client[DATABASE_NAME]

    # ── 1. Find the active session ────────────────────────────────────────────
    active_session = await db.sessions.find_one({"isActive": True})
    if not active_session:
        print("[ERROR] No active academic session found in the database.")
        client.close()
        return

    session_id  = str(active_session["_id"])
    session_name = active_session.get("name", "Unknown Session")
    print(f"[OK] Active session: {session_name}  (id: {session_id})")

    # ── 2. Find the basic dues payment for this session ───────────────────────
    # Match: category = "Dues" (case-insensitive) AND mandatory = True, active session
    dues_payments = await db.payments.find({
        "sessionId": session_id,
        "$or": [
            {"category": {"$regex": "^dues$", "$options": "i"}},
            {"title":    {"$regex": "basic due",  "$options": "i"}},
        ],
    }).to_list(length=None)

    if not dues_payments:
        print("[WARN] No 'Dues' payment found for the active session. Sheet 2 will be empty.")
        paid_user_ids: set[str] = set()
        dues_info_map: dict[str, dict] = {}
    else:
        # Collect all user IDs that appear in any dues payment's paidBy list
        paid_user_ids: set[str] = set()
        dues_info_map: dict[str, dict] = {}   # user_id -> payment info dict

        for payment in dues_payments:
            pmt_id    = str(payment["_id"])
            pmt_title = payment.get("title", "")
            pmt_amt   = payment.get("amount", 0)

            for uid in payment.get("paidBy", []):
                uid_str = str(uid)
                paid_user_ids.add(uid_str)
                if uid_str not in dues_info_map:
                    dues_info_map[uid_str] = {
                        "payment_title": pmt_title,
                        "amount":        pmt_amt,
                        "payment_date":  "",  # will be filled from transactions if available
                    }

        # Optionally enrich with actual transaction date (best-effort)
        if paid_user_ids:
            txns = await db.transactions.find({
                "sessionId": session_id,
                "paymentId": {"$in": [str(p["_id"]) for p in dues_payments]},
                "status":    "confirmed",
                "studentId": {"$in": list(paid_user_ids)},
            }).to_list(length=None)

            for txn in txns:
                uid_str = str(txn.get("studentId", ""))
                if uid_str in dues_info_map and not dues_info_map[uid_str]["payment_date"]:
                    dues_info_map[uid_str]["payment_date"] = _fmt_date(txn.get("createdAt"))

        print(f"[OK] Found {len(dues_payments)} dues payment(s) - {len(paid_user_ids)} unique payer(s)")

    # ── 3. Fetch ALL students ─────────────────────────────────────────────────
    all_users = await db.users.find(
        {"role": {"$in": ["student", "exco"]}}   # include exco (they're students too)
    ).sort([("lastName", 1), ("firstName", 1)]).to_list(length=None)

    print(f"[OK] Total students fetched: {len(all_users)}")

    # ── 4. Build the workbook ─────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── Sheet 1: All Students ─────────────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = "All Students"
    ws_all.freeze_panes = "A2"

    _header_fill(HEADER_FILL_ALL, ws_all, ALL_HEADERS)

    for row_idx, user in enumerate(all_users, start=2):
        _write_row(ws_all, row_idx, _flatten_user(user), alternate=(row_idx % 2 == 0))

    _autofit_columns(ws_all, ALL_HEADERS)

    # ── Sheet 2: Dues Paid ────────────────────────────────────────────────────
    ws_paid = wb.create_sheet(title="Dues Paid")
    ws_paid.freeze_panes = "A2"

    _header_fill(HEADER_FILL_PAID, ws_paid, PAID_HEADERS)

    paid_users = [u for u in all_users if str(u["_id"]) in paid_user_ids]
    print(f"[OK] Students who have paid dues: {len(paid_users)}")

    for row_idx, user in enumerate(paid_users, start=2):
        uid = str(user["_id"])
        info = dues_info_map.get(uid, {})
        row_data = [
            uid,
            user.get("firstName", ""),
            user.get("lastName", ""),
            user.get("matricNumber") or "",
            user.get("email", ""),
            user.get("phone") or "",
            user.get("role", ""),
            user.get("department", ""),
            str(user.get("admissionYear") or ""),
            user.get("currentLevel") or "",
            user.get("gender") or "",
            "Yes" if user.get("isActive", True) else "No",
            _fmt_date(user.get("createdAt")),
            info.get("payment_title", ""),
            info.get("amount", ""),
            info.get("payment_date", ""),
        ]
        _write_row(ws_paid, row_idx, row_data, alternate=(row_idx % 2 == 0))

    _autofit_columns(ws_paid, PAID_HEADERS)

    # ── 5. Add a summary sheet ────────────────────────────────────────────────
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.column_dimensions["A"].width = 32
    ws_summary.column_dimensions["B"].width = 24

    summary_data = [
        ("Report generated at",     datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Active session",           session_name),
        ("Total students",           len(all_users)),
        ("Students who paid dues",   len(paid_users)),
        ("Students who have NOT paid", len(all_users) - len(paid_users)),
        ("Payment rate",             f"{(len(paid_users)/len(all_users)*100):.1f}%" if all_users else "N/A"),
    ]

    for r, (label, value) in enumerate(summary_data, start=2):
        ws_summary.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=r, column=2, value=value)

    # ── 6. Save ───────────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"\n[DONE] Excel report saved to: {output_path}")
    print(f"    Sheet 1 - All Students : {len(all_users)} rows")
    print(f"    Sheet 2 - Dues Paid    : {len(paid_users)} rows")
    print(f"    Sheet 3 - Summary")

    client.close()


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Export IESA student lists to Excel (.xlsx)"
    )
    parser.add_argument(
        "--output", "-o",
        default=f"iesa_students_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        help="Output file path (default: iesa_students_<timestamp>.xlsx)",
    )
    args = parser.parse_args()

    output_path = str(Path(args.output).resolve())
    print()
    print("=" * 58)
    print("   IESA Student Excel Export")
    print("=" * 58)
    print(f"   Output : {output_path}")
    print()

    asyncio.run(run_export(output_path))


if __name__ == "__main__":
    main()
